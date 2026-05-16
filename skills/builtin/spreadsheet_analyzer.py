"""
Aethera AI - Spreadsheet Analyzer Skill

Read, write, and analyze spreadsheets (CSV and XLSX).
Supports summary statistics, row filtering, column aggregation,
anomaly detection, and data export.
"""

import csv
import io
import math
import os
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from skills.skill_base import AetheraSkill, SkillResult, skill


@skill(name="spreadsheet_analyzer", category="data")
class SpreadsheetAnalyzerSkill(AetheraSkill):
    """
    Analyze spreadsheet data from CSV and XLSX files.
    Provides summary statistics, filtering, aggregation, anomaly detection, and export.
    """

    @property
    def name(self) -> str:
        return "spreadsheet_analyzer"

    @property
    def description(self) -> str:
        return (
            "Read, analyze, filter, aggregate, and export spreadsheet data "
            "from CSV and XLSX files with summary statistics and anomaly detection"
        )

    @property
    def parameters(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": [
                        "load",
                        "summary",
                        "filter",
                        "aggregate",
                        "detect_anomalies",
                        "export",
                    ],
                    "description": "Action to perform on the spreadsheet data",
                },
                "file_path": {
                    "type": "string",
                    "description": "Path to the CSV or XLSX file",
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name for XLSX files (defaults to first sheet)",
                },
                "data": {
                    "type": "array",
                    "items": {"type": "object"},
                    "description": "Inline row data as list of dicts (alternative to file_path)",
                },
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column names to include (defaults to all)",
                },
                "filter_column": {
                    "type": "string",
                    "description": "Column name to filter on",
                },
                "filter_operator": {
                    "type": "string",
                    "enum": [
                        "eq",
                        "neq",
                        "gt",
                        "gte",
                        "lt",
                        "lte",
                        "contains",
                        "not_contains",
                        "startswith",
                        "endswith",
                        "in",
                        "not_in",
                    ],
                    "description": "Comparison operator for filtering",
                },
                "filter_value": {
                    "description": "Value to compare against for filtering",
                },
                "agg_function": {
                    "type": "string",
                    "enum": ["sum", "mean", "median", "min", "max", "count", "stdev", "variance"],
                    "description": "Aggregation function to apply",
                },
                "agg_column": {
                    "type": "string",
                    "description": "Column to aggregate",
                },
                "group_by": {
                    "type": "string",
                    "description": "Column to group by before aggregation",
                },
                "anomaly_method": {
                    "type": "string",
                    "enum": ["zscore", "iqr"],
                    "description": "Anomaly detection method: zscore or iqr",
                    "default": "iqr",
                },
                "anomaly_threshold": {
                    "type": "number",
                    "description": "Threshold for anomaly detection (z-score cutoff or IQR multiplier)",
                    "default": 1.5,
                },
                "export_format": {
                    "type": "string",
                    "enum": ["csv", "json"],
                    "description": "Export format when action is 'export'",
                    "default": "csv",
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum number of rows to load or return",
                    "default": 10000,
                },
            },
            "required": ["action"],
        }

    @property
    def examples(self) -> list:
        return [
            {
                "input": {
                    "action": "load",
                    "file_path": "data/sales.csv",
                }
            },
            {
                "input": {
                    "action": "summary",
                    "file_path": "data/sales.csv",
                }
            },
            {
                "input": {
                    "action": "filter",
                    "file_path": "data/sales.csv",
                    "filter_column": "region",
                    "filter_operator": "eq",
                    "filter_value": "West",
                }
            },
            {
                "input": {
                    "action": "aggregate",
                    "data": [{"product": "A", "revenue": 100}, {"product": "B", "revenue": 200}],
                    "agg_function": "sum",
                    "agg_column": "revenue",
                    "group_by": "product",
                }
            },
            {
                "input": {
                    "action": "detect_anomalies",
                    "file_path": "data/metrics.csv",
                    "anomaly_method": "iqr",
                    "anomaly_threshold": 1.5,
                }
            },
        ]

    async def execute(self, **kwargs) -> SkillResult:
        action = kwargs.get("action", "")
        if not action:
            return SkillResult(success=False, error="Action is required")

        try:
            if action == "load":
                return self._load(kwargs)
            elif action == "summary":
                return self._summary(kwargs)
            elif action == "filter":
                return self._filter(kwargs)
            elif action == "aggregate":
                return self._aggregate(kwargs)
            elif action == "detect_anomalies":
                return self._detect_anomalies(kwargs)
            elif action == "export":
                return self._export(kwargs)
            else:
                return SkillResult(success=False, error=f"Unknown action: {action}")
        except Exception as e:
            return SkillResult(success=False, error=f"Spreadsheet analysis failed: {e}")

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------

    def _get_rows(self, kwargs: dict) -> Tuple[List[dict], List[str]]:
        """Resolve rows from either file_path or inline data."""
        file_path = kwargs.get("file_path")
        inline_data = kwargs.get("data")

        if file_path:
            rows, headers = self._read_file(
                file_path,
                sheet_name=kwargs.get("sheet_name"),
                limit=kwargs.get("limit", 10000),
            )
        elif inline_data:
            if not isinstance(inline_data, list) or not inline_data:
                raise ValueError("Inline data must be a non-empty list of dicts")
            rows = inline_data
            headers = list(inline_data[0].keys()) if isinstance(inline_data[0], dict) else []
        else:
            raise ValueError("Either file_path or data must be provided")

        # Column selection
        selected = kwargs.get("columns")
        if selected:
            rows = [{k: v for k, v in row.items() if k in selected} for row in rows]
            headers = [h for h in headers if h in selected]

        return rows, headers

    def _read_file(
        self, file_path: str, sheet_name: Optional[str] = None, limit: int = 10000
    ) -> Tuple[List[dict], List[str]]:
        """Read a CSV or XLSX file and return rows as dicts plus headers."""
        ext = Path(file_path).suffix.lower()

        if ext == ".csv":
            return self._read_csv(file_path, limit)
        elif ext in (".xlsx", ".xls"):
            return self._read_xlsx(file_path, sheet_name, limit)
        else:
            # Attempt CSV as fallback
            try:
                return self._read_csv(file_path, limit)
            except Exception:
                raise ValueError(f"Unsupported file format: {ext}. Use .csv or .xlsx")

    def _read_csv(self, file_path: str, limit: int = 10000) -> Tuple[List[dict], List[str]]:
        """Read a CSV file."""
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        rows: List[dict] = []
        headers: List[str] = []

        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            headers = list(reader.fieldnames or [])
            for i, row in enumerate(reader):
                if i >= limit:
                    break
                # Cast numeric strings
                rows.append(self._cast_row(row))

        return rows, headers

    def _read_xlsx(
        self, file_path: str, sheet_name: Optional[str] = None, limit: int = 10000
    ) -> Tuple[List[dict], List[str]]:
        """Read an XLSX file using openpyxl."""
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            from openpyxl import load_workbook
        except ImportError:
            # Fallback: try reading as CSV if openpyxl is not installed
            return self._read_csv(file_path, limit)

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        if ws is None:
            wb.close()
            raise ValueError(f"Sheet not found: {sheet_name}")

        rows_iter = ws.iter_rows(values_only=True)
        # First row is headers
        try:
            header_tuple = next(rows_iter)
        except StopIteration:
            wb.close()
            return [], []

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_tuple)]
        rows: List[dict] = []

        for i, values in enumerate(rows_iter):
            if i >= limit:
                break
            row_dict = {}
            for j, val in enumerate(values):
                key = headers[j] if j < len(headers) else f"col_{j}"
                row_dict[key] = val
            rows.append(self._cast_row(row_dict))

        wb.close()
        return rows, headers

    @staticmethod
    def _cast_row(row: dict) -> dict:
        """Attempt to cast string values to int/float where possible."""
        casted = {}
        for k, v in row.items():
            if v is None:
                casted[k] = None
                continue
            if isinstance(v, (int, float)):
                casted[k] = v
                continue
            s = str(v).strip()
            # Try int first, then float
            try:
                casted[k] = int(s)
                continue
            except (ValueError, TypeError):
                pass
            try:
                casted[k] = float(s)
                continue
            except (ValueError, TypeError):
                pass
            casted[k] = s
        return casted

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def _load(self, kwargs: dict) -> SkillResult:
        """Load and preview spreadsheet data."""
        rows, headers = self._get_rows(kwargs)
        limit = kwargs.get("limit", 10000)
        preview_count = min(20, len(rows))

        return SkillResult(
            success=True,
            data={
                "total_rows": len(rows),
                "total_columns": len(headers),
                "headers": headers,
                "preview": rows[:preview_count],
                "preview_count": preview_count,
                "loaded_count": len(rows),
                "truncated": len(rows) >= limit,
            },
        )

    def _summary(self, kwargs: dict) -> SkillResult:
        """Compute summary statistics for each column."""
        rows, headers = self._get_rows(kwargs)

        if not rows:
            return SkillResult(success=True, data={"columns": {}, "total_rows": 0})

        column_stats: Dict[str, Any] = {}

        for col in headers:
            values = [row.get(col) for row in rows if row.get(col) is not None]
            numeric_values = [v for v in values if isinstance(v, (int, float))]

            stats: Dict[str, Any] = {
                "count": len(values),
                "null_count": len(rows) - len(values),
            }

            if numeric_values:
                stats["type"] = "numeric"
                stats["min"] = min(numeric_values)
                stats["max"] = max(numeric_values)
                stats["mean"] = round(statistics.mean(numeric_values), 4)
                stats["median"] = statistics.median(numeric_values)
                if len(numeric_values) > 1:
                    stats["stdev"] = round(statistics.stdev(numeric_values), 4)
                    stats["variance"] = round(statistics.variance(numeric_values), 4)
                stats["sum"] = sum(numeric_values)
                # Quartiles
                sorted_vals = sorted(numeric_values)
                n = len(sorted_vals)
                q1_idx = n // 4
                q3_idx = 3 * n // 4
                stats["q1"] = sorted_vals[q1_idx]
                stats["q3"] = sorted_vals[q3_idx]
                stats["iqr"] = stats["q3"] - stats["q1"]
            else:
                stats["type"] = "categorical"
                # Frequency distribution for categoricals
                freq: Dict[str, int] = {}
                for v in values:
                    key = str(v)
                    freq[key] = freq.get(key, 0) + 1
                sorted_freq = sorted(freq.items(), key=lambda x: x[1], reverse=True)
                stats["unique_count"] = len(freq)
                stats["top_values"] = sorted_freq[:10]
                if len(values) > 0:
                    stats["mode"] = sorted_freq[0][0] if sorted_freq else None

            column_stats[col] = stats

        return SkillResult(
            success=True,
            data={
                "total_rows": len(rows),
                "total_columns": len(headers),
                "columns": column_stats,
            },
        )

    def _filter(self, kwargs: dict) -> SkillResult:
        """Filter rows based on column value comparison."""
        rows, headers = self._get_rows(kwargs)
        filter_col = kwargs.get("filter_column")
        filter_op = kwargs.get("filter_operator", "eq")
        filter_val = kwargs.get("filter_value")

        if not filter_col:
            return SkillResult(success=False, error="filter_column is required for filter action")
        if filter_val is None:
            return SkillResult(success=False, error="filter_value is required for filter action")
        if filter_col not in headers:
            return SkillResult(success=False, error=f"Column '{filter_col}' not found. Available: {headers}")

        # If filter_value is a string list representation for 'in'/'not_in', split it
        if filter_op in ("in", "not_in") and isinstance(filter_val, str):
            filter_val = [v.strip() for v in filter_val.split(",")]

        operators = {
            "eq": lambda a, b: a == b,
            "neq": lambda a, b: a != b,
            "gt": lambda a, b: self._safe_numeric_compare(a, b, lambda x, y: x > y),
            "gte": lambda a, b: self._safe_numeric_compare(a, b, lambda x, y: x >= y),
            "lt": lambda a, b: self._safe_numeric_compare(a, b, lambda x, y: x < y),
            "lte": lambda a, b: self._safe_numeric_compare(a, b, lambda x, y: x <= y),
            "contains": lambda a, b: str(b) in str(a),
            "not_contains": lambda a, b: str(b) not in str(a),
            "startswith": lambda a, b: str(a).startswith(str(b)),
            "endswith": lambda a, b: str(a).endswith(str(b)),
            "in": lambda a, b: a in b,
            "not_in": lambda a, b: a not in b,
        }

        op_func = operators.get(filter_op)
        if op_func is None:
            return SkillResult(success=False, error=f"Unknown operator: {filter_op}")

        # Cast filter_val for numeric comparisons
        if filter_op in ("gt", "gte", "lt", "lte") and isinstance(filter_val, str):
            try:
                filter_val = float(filter_val)
            except ValueError:
                pass

        filtered = []
        for row in rows:
            cell_val = row.get(filter_col)
            if cell_val is None:
                continue
            try:
                if op_func(cell_val, filter_val):
                    filtered.append(row)
            except (TypeError, ValueError):
                continue

        return SkillResult(
            success=True,
            data={
                "filter": {
                    "column": filter_col,
                    "operator": filter_op,
                    "value": filter_val,
                },
                "total_rows": len(rows),
                "matched_rows": len(filtered),
                "rows": filtered,
            },
        )

    @staticmethod
    def _safe_numeric_compare(a: Any, b: Any, op) -> bool:
        """Attempt numeric comparison, return False on failure."""
        try:
            a_num = float(a) if not isinstance(a, (int, float)) else a
            b_num = float(b) if not isinstance(b, (int, float)) else b
            return op(a_num, b_num)
        except (TypeError, ValueError):
            return False

    def _aggregate(self, kwargs: dict) -> SkillResult:
        """Aggregate column values, optionally grouped by another column."""
        rows, headers = self._get_rows(kwargs)
        agg_func = kwargs.get("agg_function")
        agg_col = kwargs.get("agg_column")
        group_by = kwargs.get("group_by")

        if not agg_func:
            return SkillResult(success=False, error="agg_function is required for aggregate action")
        if not agg_col:
            return SkillResult(success=False, error="agg_column is required for aggregate action")
        if agg_col not in headers:
            return SkillResult(success=False, error=f"Column '{agg_col}' not found. Available: {headers}")
        if group_by and group_by not in headers:
            return SkillResult(success=False, error=f"Group-by column '{group_by}' not found. Available: {headers}")

        if group_by:
            groups: Dict[str, List[Any]] = {}
            for row in rows:
                key = str(row.get(group_by, ""))
                val = row.get(agg_col)
                if val is not None:
                    groups.setdefault(key, []).append(val)

            result = {}
            for key, values in sorted(groups.items()):
                result[key] = self._apply_agg(agg_func, values)
        else:
            values = [row.get(agg_col) for row in rows if row.get(agg_col) is not None]
            result = self._apply_agg(agg_func, values)

        return SkillResult(
            success=True,
            data={
                "aggregation": {
                    "function": agg_func,
                    "column": agg_col,
                    "group_by": group_by,
                },
                "result": result,
                "total_rows": len(rows),
            },
        )

    @staticmethod
    def _apply_agg(func_name: str, values: list) -> Any:
        """Apply an aggregation function to a list of values."""
        numeric = [v for v in values if isinstance(v, (int, float))]

        if func_name == "count":
            return len(values)
        if func_name == "sum":
            return sum(numeric) if numeric else 0
        if func_name == "mean":
            return round(statistics.mean(numeric), 4) if numeric else None
        if func_name == "median":
            return statistics.median(numeric) if numeric else None
        if func_name == "min":
            return min(numeric) if numeric else None
        if func_name == "max":
            return max(numeric) if numeric else None
        if func_name == "stdev":
            return round(statistics.stdev(numeric), 4) if len(numeric) > 1 else None
        if func_name == "variance":
            return round(statistics.variance(numeric), 4) if len(numeric) > 1 else None
        raise ValueError(f"Unknown aggregation function: {func_name}")

    def _detect_anomalies(self, kwargs: dict) -> SkillResult:
        """Detect anomalous values in numeric columns."""
        rows, headers = self._get_rows(kwargs)
        method = kwargs.get("anomaly_method", "iqr")
        threshold = kwargs.get("anomaly_threshold", 1.5)

        anomalies: Dict[str, Any] = {}
        total_anomaly_count = 0

        # Identify numeric columns
        numeric_cols = []
        for col in headers:
            numeric_vals = [row.get(col) for row in rows if isinstance(row.get(col), (int, float))]
            if len(numeric_vals) > 3:
                numeric_cols.append(col)

        if not numeric_cols:
            return SkillResult(
                success=True,
                data={"anomalies": {}, "total_anomalies": 0, "message": "No numeric columns with sufficient data for anomaly detection"},
            )

        for col in numeric_cols:
            values = [row.get(col) for row in rows if isinstance(row.get(col), (int, float))]
            if len(values) < 4:
                continue

            mean = statistics.mean(values)
            stdev = statistics.stdev(values) if len(values) > 1 else 0
            sorted_vals = sorted(values)
            n = len(sorted_vals)
            q1 = sorted_vals[n // 4]
            q3 = sorted_vals[3 * n // 4]
            iqr = q3 - q1

            col_anomalies: List[Dict[str, Any]] = []

            for row_idx, row in enumerate(rows):
                val = row.get(col)
                if not isinstance(val, (int, float)):
                    continue

                is_anomaly = False
                reason = ""

                if method == "zscore" and stdev > 0:
                    z = abs(val - mean) / stdev
                    if z > threshold:
                        is_anomaly = True
                        reason = f"z-score {z:.2f} exceeds threshold {threshold}"
                elif method == "iqr":
                    lower = q1 - threshold * iqr
                    upper = q3 + threshold * iqr
                    if val < lower or val > upper:
                        is_anomaly = True
                        bound = "below lower" if val < lower else "above upper"
                        reason = f"value {val} is {bound} bound ({lower:.2f}, {upper:.2f})"

                if is_anomaly:
                    col_anomalies.append({
                        "row_index": row_idx,
                        "value": val,
                        "reason": reason,
                    })
                    total_anomaly_count += 1

            anomalies[col] = {
                "count": len(col_anomalies),
                "mean": round(mean, 4),
                "stdev": round(stdev, 4),
                "q1": q1,
                "q3": q3,
                "iqr": iqr,
                "outliers": col_anomalies[:100],  # Cap at 100 for response size
            }

        return SkillResult(
            success=True,
            data={
                "method": method,
                "threshold": threshold,
                "columns_analyzed": len(numeric_cols),
                "total_anomalies": total_anomaly_count,
                "anomalies": anomalies,
                "total_rows": len(rows),
            },
        )

    def _export(self, kwargs: dict) -> SkillResult:
        """Export data to CSV or JSON format."""
        rows, headers = self._get_rows(kwargs)
        export_format = kwargs.get("export_format", "csv")

        if export_format == "csv":
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
            content = output.getvalue()
            output.close()

            return SkillResult(
                success=True,
                data={
                    "format": "csv",
                    "content": content,
                    "row_count": len(rows),
                    "column_count": len(headers),
                },
            )
        elif export_format == "json":
            return SkillResult(
                success=True,
                data={
                    "format": "json",
                    "rows": rows,
                    "headers": headers,
                    "row_count": len(rows),
                    "column_count": len(headers),
                },
            )
        else:
            return SkillResult(success=False, error=f"Unsupported export format: {export_format}")