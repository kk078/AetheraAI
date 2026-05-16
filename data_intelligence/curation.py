"""
Aethera AI — Stage 1: Dataset Curation

Collects, cleans, deduplicates, and validates data from any source.
Loads CSV/XLSX/JSON files or inline data, normalizes null values,
removes duplicates, and validates required columns.
"""

import csv
import hashlib
import io
import json
import logging
from typing import Any, Dict, List, Optional, Set

from .context import DataPipelineContext
from .stages import DataPipelineStage

logger = logging.getLogger("aethera.data_intelligence.curation")

# Values that should be treated as null
NULL_VALUES = {"", "na", "n/a", "null", "none", "nil", "-", "--", "undefined", "nan", "NULL", "None"}


class DatasetCurationStage(DataPipelineStage):
    """Load, clean, deduplicate, and validate data."""

    name = "curation"

    @property
    def required_content(self) -> bool:
        return False  # This stage loads data, doesn't require pre-existing content

    async def execute(self, context: DataPipelineContext) -> DataPipelineContext:
        # Load data
        rows, headers = await self._load_data(context)
        if not rows:
            context.validation_errors.append("No data loaded")
            return context

        context.rows = rows
        context.headers = headers

        # Clean
        cleaned = self._clean_rows(rows, headers)
        context.cleaned_rows = cleaned

        # Deduplicate
        deduped, removed = self._deduplicate(cleaned)
        context.cleaned_rows = deduped
        context.duplicates_removed = removed

        # Validate
        errors = self._validate(deduped, headers, context.options)
        context.validation_errors = errors

        # Update dataset metadata
        context = self._update_metadata(context)

        logger.info(
            f"Curation: {len(rows)} rows loaded, "
            f"{removed} duplicates removed, "
            f"{len(deduped)} rows remaining, "
            f"{len(errors)} validation issues"
        )
        return context

    async def _load_data(
        self, context: DataPipelineContext
    ) -> tuple:
        """Load data from file or inline source."""
        source_type = context.source_type
        source_path = context.source_path

        if source_type == "inline":
            # Data provided directly in context
            rows = context.options.get("data", [])
            if not rows and context.rows:
                rows = context.rows
            if rows:
                headers = list(rows[0].keys()) if rows else []
                return rows, headers
            return [], []

        if not source_path:
            # Try to use rows already in context
            if context.rows:
                headers = context.headers or list(context.rows[0].keys())
                return context.rows, headers
            return [], []

        fmt = context.format or self._detect_format(source_path)

        try:
            if fmt == "csv":
                return self._load_csv(source_path)
            elif fmt == "json":
                return self._load_json(source_path)
            elif fmt == "xlsx":
                return self._load_xlsx(source_path)
            else:
                logger.warning(f"Unsupported format: {fmt}, attempting CSV")
                return self._load_csv(source_path)
        except Exception as e:
            logger.error(f"Failed to load data from {source_path}: {e}")
            return [], []

    def _detect_format(self, path: str) -> str:
        """Detect file format from extension."""
        path_lower = path.lower()
        if path_lower.endswith(".csv"):
            return "csv"
        elif path_lower.endswith(".json") or path_lower.endswith(".jsonl"):
            return "json"
        elif path_lower.endswith(".xlsx"):
            return "xlsx"
        elif path_lower.endswith(".parquet"):
            return "parquet"
        return "csv"  # default

    def _load_csv(self, path: str) -> tuple:
        """Load data from a CSV file."""
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            rows = [dict(row) for row in reader]
            headers = list(reader.fieldnames or [])
        return rows, headers

    def _load_json(self, path: str) -> tuple:
        """Load data from a JSON file (array of objects or newline-delimited)."""
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()

        try:
            data = json.loads(content)
            if isinstance(data, list) and len(data) > 0:
                headers = list(data[0].keys())
                return data, headers
        except json.JSONDecodeError:
            # Try newline-delimited JSON
            rows = []
            for line in content.strip().split("\n"):
                line = line.strip()
                if line:
                    try:
                        rows.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
            if rows:
                headers = list(rows[0].keys())
                return rows, headers
        return [], []

    def _load_xlsx(self, path: str) -> tuple:
        """Load data from an XLSX file using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed, cannot read XLSX")
            return [], []

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        rows_data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell is not None else f"column_{i}" for i, cell in enumerate(row)]
            else:
                row_dict = {}
                for j, cell in enumerate(row):
                    key = headers[j] if j < len(headers) else f"column_{j}"
                    row_dict[key] = cell
                rows_data.append(row_dict)

        wb.close()
        return rows_data, headers

    def _clean_rows(
        self, rows: List[Dict[str, Any]], headers: List[str]
    ) -> List[Dict[str, Any]]:
        """Clean row data: normalize nulls, strip whitespace, type-cast values."""
        cleaned = []
        for row in rows:
            cleaned_row = {}
            for key in headers:
                value = row.get(key)
                cleaned_row[key] = self._clean_value(value)
            cleaned.append(cleaned_row)
        return cleaned

    def _clean_value(self, value: Any) -> Any:
        """Clean a single value: normalize nulls and strip strings."""
        if value is None:
            return None

        if isinstance(value, str):
            stripped = value.strip()
            if stripped.lower() in NULL_VALUES:
                return None
            # Try numeric conversion
            try:
                if "." in stripped:
                    return float(stripped)
                return int(stripped)
            except ValueError:
                return stripped

        return value

    def _deduplicate(
        self, rows: List[Dict[str, Any]]
    ) -> tuple:
        """Remove exact duplicate rows based on content hash."""
        seen: Set[str] = set()
        deduped = []
        removed = 0

        for row in rows:
            row_hash = hashlib.md5(
                json.dumps(row, sort_keys=True, default=str).encode()
            ).hexdigest()

            if row_hash in seen:
                removed += 1
                continue

            seen.add(row_hash)
            deduped.append(row)

        return deduped, removed

    def _validate(
        self,
        rows: List[Dict[str, Any]],
        headers: List[str],
        options: Dict[str, Any],
    ) -> List[str]:
        """Validate data against constraints specified in options."""
        errors = []

        if not rows:
            errors.append("Dataset is empty after cleaning")
            return errors

        # Check for empty headers
        empty_headers = [h for h in headers if not h.strip()]
        if empty_headers:
            errors.append(f"Empty column headers found at positions: {[headers.index(h) for h in empty_headers]}")

        # Check for required columns (if specified in options)
        required = options.get("required_columns", [])
        for col in required:
            if col not in headers:
                errors.append(f"Required column missing: {col}")

        # Check for columns that are entirely null
        null_counts = {h: 0 for h in headers}
        for row in rows:
            for h in headers:
                if row.get(h) is None:
                    null_counts[h] += 1

        total = len(rows)
        for h, count in null_counts.items():
            if count == total:
                errors.append(f"Column '{h}' is entirely null")

        return errors

    def _update_metadata(self, context: DataPipelineContext) -> DataPipelineContext:
        """Update context metadata after curation."""
        rows = context.cleaned_rows
        if rows:
            context.headers = list(rows[0].keys())
        return context