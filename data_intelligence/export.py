"""
Aethera AI — Stage 7: Data Export

Export datasets in CSV, JSON, Parquet, or XLSX format.
Parquet requires pyarrow; XLSX requires openpyxl.
"""

import csv
import io
import json
import logging
from typing import Any, Dict, List, Optional

from .context import DataPipelineContext
from .stages import DataPipelineStage

logger = logging.getLogger("aethera.data_intelligence.export")


class DataExportStage(DataPipelineStage):
    """Export dataset in the requested format."""

    name = "export"

    @property
    def required_content(self) -> bool:
        return True

    async def execute(self, context: DataPipelineContext) -> DataPipelineContext:
        rows = context.cleaned_rows or context.rows
        if not rows:
            logger.warning("Export: no rows to export")
            return context

        export_format = context.options.get("export_format", context.export_format or "csv")
        columns = context.options.get("export_columns", None)
        output_path = context.options.get("output_path", None)

        # Filter columns if specified
        if columns:
            filtered_rows = []
            for row in rows:
                filtered_rows.append({k: v for k, v in row.items() if k in columns})
            rows = filtered_rows

        # Export based on format
        if export_format == "csv":
            content, content_bytes = self._export_csv(rows)
        elif export_format == "json":
            content, content_bytes = self._export_json(rows)
        elif export_format == "jsonl":
            content, content_bytes = self._export_jsonl(rows)
        elif export_format == "parquet":
            content, content_bytes = self._export_parquet(rows)
        elif export_format == "xlsx":
            content, content_bytes = self._export_xlsx(rows)
        else:
            logger.warning(f"Unknown export format: {export_format}, defaulting to CSV")
            content, content_bytes = self._export_csv(rows)
            export_format = "csv"

        # Write to file if output_path specified
        if output_path:
            import os
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            if content_bytes:
                with open(output_path, "wb") as f:
                    f.write(content_bytes)
            else:
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(content)
            context.export_path = output_path

        context.export_content = content
        context.export_bytes = content_bytes
        context.export_format = export_format

        logger.info(f"Export: {len(rows)} rows in {export_format} format")
        return context

    def _export_csv(self, rows: List[Dict[str, Any]]) -> tuple:
        """Export as CSV. Returns (string_content, None)."""
        if not rows:
            return "", None

        headers = list(rows[0].keys())
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

        return output.getvalue(), None

    def _export_json(self, rows: List[Dict[str, Any]]) -> tuple:
        """Export as JSON array. Returns (string_content, None)."""
        content = json.dumps(rows, indent=2, default=str)
        return content, None

    def _export_jsonl(self, rows: List[Dict[str, Any]]) -> tuple:
        """Export as newline-delimited JSON. Returns (string_content, None)."""
        lines = [json.dumps(row, default=str) for row in rows]
        content = "\n".join(lines)
        return content, None

    def _export_parquet(self, rows: List[Dict[str, Any]]) -> tuple:
        """Export as Parquet. Requires pyarrow. Returns (None, bytes_content)."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            logger.error("pyarrow not installed. Install with: pip install pyarrow")
            return None, None

        if not rows:
            return None, None

        # Build PyArrow table from rows
        headers = list(rows[0].keys())
        columns = {h: [] for h in headers}

        for row in rows:
            for h in headers:
                columns[h].append(row.get(h))

        # Infer types and create arrays
        arrays = []
        fields = []
        for h in headers:
            values = columns[h]
            # Try to convert to native types
            try:
                arrays.append(pa.array(values))
                fields.append(pa.field(h, pa.array(values).type))
            except Exception:
                # Fall back to string
                str_values = [str(v) if v is not None else None for v in values]
                arrays.append(pa.array(str_values))
                fields.append(pa.field(h, pa.string()))

        schema = pa.schema(fields)
        table = pa.table({h: a for h, a in zip(headers, arrays)}, schema=schema)

        # Write to buffer
        buf = io.BytesIO()
        pq.write_table(table, buf)
        buf.seek(0)

        return None, buf.read()

    def _export_xlsx(self, rows: List[Dict[str, Any]]) -> tuple:
        """Export as XLSX. Requires openpyxl. Returns (None, bytes_content)."""
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return None, None

        if not rows:
            return None, None

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Write headers
        headers = list(rows[0].keys())
        ws.append(headers)

        # Write data rows
        for row in rows:
            ws.append([row.get(h) for h in headers])

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=min(len(rows) + 1, 100)):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return None, buf.read()